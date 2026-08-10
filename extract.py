import openpyxl, re, csv, unicodedata

def norm(s):
    if s is None: return ""
    s = str(s).strip()
    s2 = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode('ascii')
    s2 = s2.replace('-', ' ')
    s2 = re.sub(r'\s+', ' ', s2)
    return s2.lower().strip()

def cle_arret(commune, arret):
    """Clé unique d'un arrêt = commune + nom d'arrêt normalisés.
    Nécessaire car plusieurs arrêts de communes différentes portent le
    même nom (ex: 'Marché' existe à Saint-Trojan ET à Saint-Pierre)."""
    return f"{norm(commune)}|{norm(arret)}"

# ---------- 1. Arrets.xlsx ----------
wb = openpyxl.load_workbook('/mnt/user-data/uploads/Arrêts.xlsx', data_only=True)
ws = wb['Feuil1']
arrets = []
for row in ws.iter_rows(min_row=2, values_only=True):
    commune, arret, pi, corresp = row[0], row[1], row[2], row[3]
    if not arret: continue
    arrets.append({
        'commune': commune,
        'arret': arret,
        'arret_norm': cle_arret(commune, arret),
        'point_interet': 1 if pi else 0,
        'correspondance': corresp or ""
    })

with open('arrets.csv','w',newline='',encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['commune','arret','arret_norm','point_interet','correspondance'])
    w.writeheader()
    vus = set()
    arrets_dedup = []
    for a in arrets:
        if a['arret_norm'] in vus:
            continue  # doublon dans Arrêts.xlsx (même arrêt saisi deux fois avec une orthographe différente)
        vus.add(a['arret_norm'])
        arrets_dedup.append(a)
    w.writerows(arrets_dedup)
print("arrets.csv:", len(arrets_dedup), "lignes (", len(arrets) - len(arrets_dedup), "doublons retirés)")

# ---------- 2. Fiches_Horaires.xlsx ----------
wb2 = openpyxl.load_workbook('/mnt/user-data/uploads/Fiches_Horaires.xlsx', data_only=True)
ws2 = wb2['Feuil1']
rows = list(ws2.iter_rows(values_only=True))

block_re = re.compile(r'^(\d+)\s+Navette\s+(\S+)\s+—\s+(.+?)\s*>\s*(.+)$')
course_re = re.compile(r'Course\s+(\d+)\s*\(([0-9]{1,2}:[0-9]{2})\)')

current = None
long_rows = []
i = 0
n = len(rows)
while i < n:
    row = rows[i]
    c0 = row[0]
    if c0 and isinstance(c0, str):
        m = block_re.match(c0.strip())
        if m:
            current = {
                'ligne_num': int(m.group(1)),
                'couleur': m.group(2),
                'origine': m.group(3).strip(),
                'destination': m.group(4).strip(),
                'jours_validite': None,
            }
            # next row usually = validity days text
            if i+1 < n and rows[i+1][0] and 'Source' not in str(rows[i+1][0]):
                current['jours_validite'] = str(rows[i+1][0]).strip()
            i += 1
            continue
    if c0 == 'Commune' and row[1] == 'Arrêt' and current is not None:
        header = row
        # find course columns and remarque column
        course_cols = {}  # col_index -> (course_num, heure_ref)
        remarque_idx = None
        for idx, val in enumerate(header):
            if val is None: continue
            cm = course_re.search(str(val))
            if cm:
                course_cols[idx] = (int(cm.group(1)), cm.group(2))
            elif str(val).strip() == 'Remarque':
                remarque_idx = idx
        j = i + 1
        while j < n and rows[j][0] is not None:
            r = rows[j]
            commune = r[0]
            arret = r[1]
            remarque = r[remarque_idx] if remarque_idx is not None else None
            for idx, (cnum, href) in course_cols.items():
                t = r[idx]
                if t is not None:
                    hhmm = t.strftime('%H:%M') if hasattr(t, 'strftime') else str(t)
                    long_rows.append({
                        'ligne_num': current['ligne_num'],
                        'couleur': current['couleur'],
                        'origine': current['origine'],
                        'destination': current['destination'],
                        'jours_validite': current['jours_validite'],
                        'course_num': cnum,
                        'heure_ref_depart_ligne': href,
                        'commune': commune,
                        'arret': arret,
                        'arret_norm': cle_arret(commune, arret),
                        'heure_passage': hhmm,
                        'remarque': remarque or ""
                    })
            j += 1
        i = j
        continue
    i += 1

with open('horaires_long.csv','w',newline='',encoding='utf-8') as f:
    fieldnames = ['ligne_num','couleur','origine','destination','jours_validite','course_num',
                  'heure_ref_depart_ligne','commune','arret','arret_norm','heure_passage','remarque']
    w = csv.DictWriter(f, fieldnames=fieldnames)
    w.writeheader()
    w.writerows(long_rows)
print("horaires_long.csv:", len(long_rows), "lignes")

# ---------- 3. controle correspondance noms d'arrets ----------
arret_norms_ref = set(a['arret_norm'] for a in arrets)
arret_norms_hor = set(r['arret_norm'] for r in long_rows)
manquants = sorted(arret_norms_hor - arret_norms_ref)
print("\nArrets presents dans horaires mais absents (apres normalisation) de Arrets.xlsx:", len(manquants))
for x in manquants:
    print(" -", x)
